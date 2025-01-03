"""
CSV Parser Module for AION License Count Application

This module is responsible for processing CSV files containing Microsoft license data
and generating detailed Excel reports with license counts and cost calculations.

Key functionalities:
1. Reading and cleaning CSV data
2. Counting licenses per office and license type
3. Categorizing users (AION Management, AION Partners, Properties, Unaccounted)
4. Calculating costs based on license counts
5. Generating a formatted Excel report with multiple sheets
6. Creating a summary of license usage and costs

The main entry point is the `process_file` function, which orchestrates the entire
process from reading the input CSV to producing the final Excel report.

This module is designed to work with the Flask web application, utilizing
configuration settings and logging mechanisms defined in other parts of the project.

Dependencies:
- pandas: for data manipulation
- openpyxl: for Excel file operations
- utils.logger: for logging
- create_app: for application configuration

Note: This module assumes a specific structure for the input CSV file, including
columns for 'Office', 'Licenses', 'User principal name', and 'Display name'.
"""

import pandas as pd
import os
import openpyxl
import time
import uuid
from datetime import datetime
from utils.logger import get_logger
from create_app import app
import os.path
from pathlib import Path

logger = get_logger(__name__)


def sanitize_path(base_path, filename):
    """
    Sanitize a file path to prevent path traversal attacks.
    
    Args:
        base_path (str): The base directory path
        filename (str): The filename to sanitize
        
    Returns:
        str: A sanitized absolute path that is guaranteed to be under base_path
    """
    base_path = os.path.abspath(base_path)
    try:
        full_path = os.path.abspath(os.path.join(base_path, filename))
        if not full_path.startswith(base_path):
            raise ValueError("Path traversal detected")
        return full_path
    except Exception as e:
        logger.error(f"Path sanitization failed: {e}")
        raise ValueError("Invalid path")


def read_and_prepare_data(file_path):
    """
    Read CSV file and prepare the DataFrame by stripping whitespaces from the 'Office' column.

    Args:
        file_path (str): Path to the CSV file.

    Returns:
        pandas.DataFrame or None: Prepared DataFrame if successful, None otherwise.
    """
    try:
        df = pd.read_csv(file_path)
        df['Office'] = df['Office'].str.strip()
        logger.info(f"csv cleaned successfully from {file_path}")
        return df
    except FileNotFoundError:
        logger.error(f"File not found: {file_path}")
        return None
    except pd.errors.EmptyDataError:
        logger.error(f"No data: {file_path} is empty")
        return None
    except Exception as e:
        logger.error(f"Error reading file {file_path}: {e}")
        return None


def initialize_license_counts(df, target_licenses):
    """
    Initialize the license counts dictionary based on unique 'Office' values and target licenses.

    Args:
        df (pandas.DataFrame): The input DataFrame.
        target_licenses (dict): Dictionary of target license types.

    Returns:
        dict: Initialized license counts dictionary.
    """
    license_counts = {office.strip(): {key: 0 for key in target_licenses.keys()}
                      for office in df['Office'].unique() if pd.notna(office) and office.strip()}
    license_counts['Unaccounted'] = {key: 0 for key in target_licenses.keys()}
    logger.info("Initialized license counts")
    return license_counts


def process_licenses(df, target_licenses, license_counts):
    """
    Process each row in the DataFrame to count licenses and log specific organizations.

    Args:
        df (pandas.DataFrame): The input DataFrame.
        target_licenses (dict): Dictionary of target license types.
        license_counts (dict): Initialized license counts dictionary.

    Returns:
        tuple: Updated license_counts, unaccounted_users, aion_management, aion_partners, properties
    """
    logger.info("Processing licenses")
    unaccounted_users = []
    aion_management = []
    aion_partners = []
    properties = []

    for index, row in df.iterrows():
        office = row.get('Office')
        licenses = row.get('Licenses')
        user_principal_name = row.get('User principal name')
        display_name = row.get('Display name')

        if pd.isna(licenses) or licenses.strip() == '':
            continue

        licenses_list = licenses.split('+')

        for license in licenses_list:
            license = license.strip()

            for key, variants in target_licenses.items():
                if any(variant in license for variant in variants):
                    license_counts[office if not (pd.isna(office) or office.strip() == '') else 'Unaccounted'][key] += 1
                    if office == 'AION Management':
                        aion_management.append([display_name, license, user_principal_name])
                    elif office == 'AION Partners':
                        aion_partners.append([display_name, license, user_principal_name])
                    elif pd.isna(office) or office.strip() == '':
                        unaccounted_users.append([display_name, license, user_principal_name])
                    else:
                        properties.append([display_name, license, user_principal_name, office])

    logger.info("Processed licenses")
    return license_counts, unaccounted_users, aion_management, aion_partners, properties


def create_license_counts_df(license_counts):
    """
    Convert the license counts dictionary to a DataFrame and calculate totals.

    Args:
        license_counts (dict): License counts dictionary.

    Returns:
        pandas.DataFrame or None: DataFrame with license counts and totals if successful, None otherwise.
    """
    try:
        license_counts_df = pd.DataFrame.from_dict(license_counts, orient='index').fillna(0)
        totals = license_counts_df.sum(axis=0).rename('Total')
        license_counts_df = pd.concat([license_counts_df, pd.DataFrame(totals).T])
        logger.info("Created license counts DataFrame")
        return license_counts_df
    except Exception as e:
        logger.error(f"Error creating license counts DataFrame: {e}")
        return None


def save_to_excel(excel_path, license_counts_df, aion_management_df, aion_partners_df, properties_df, unaccounted_users,
                  cost_per_user, cost_per_exchange, cost_per_e5, cost_per_teams):
    """
       Save the processed data to an Excel file with specific formatting.

       Args:
           excel_path (str): Path to save the Excel file.
           license_counts_df (pandas.DataFrame): DataFrame with license counts.
           aion_management_df (pandas.DataFrame): DataFrame with AION Management data.
           aion_partners_df (pandas.DataFrame): DataFrame with AION Partners data.
           properties_df (pandas.DataFrame): DataFrame with Properties data.
           unaccounted_users (pandas.DataFrame): DataFrame with unaccounted users' data.
           cost_per_user (int): Cost per user.
           cost_per_exchange (int): Cost per exchange license.
           cost_per_e5 (int): Cost per E5 license.
           cost_per_teams (int): Cost per Teams license.
       """
    try:
        # Sanitize output path
        base_dir = os.path.dirname(os.path.abspath(excel_path))
        excel_path = sanitize_path(base_dir, os.path.basename(excel_path))
        
        logger.info(f"writing data to Excel file: {excel_path}")
        with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
            license_counts_df.to_excel(writer, sheet_name='License Counts')
            aion_management_df.to_excel(writer, sheet_name='AION Management', index=False)
            aion_partners_df.to_excel(writer, sheet_name='AION Partners', index=False)
            properties_df.to_excel(writer, sheet_name='Properties', index=False)
            unaccounted_users.to_excel(writer, sheet_name='Unaccounted Users', index=False)

            workbook = writer.book
            header_format = workbook.add_format(
                {'bold': True, 'text_wrap': True, 'valign': 'top', 'fg_color': '#D7E4BC', 'border': 1})
            total_format = workbook.add_format({'bold': True, 'fg_color': '#FFEB9C', 'border': 1,
                                                'num_format': '#,##0'})
            currency_format = workbook.add_format({'num_format': '$#,##0'})

            def format_sheet(worksheet, df, is_totals=False):
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num + 1 if is_totals else col_num, value, header_format)
                if is_totals:
                    worksheet.write(0, 0, 'Office', header_format)

                for col_num, col in enumerate(df.columns):
                    column_len = df[col].astype(str).str.len().max()
                    column_len = max(column_len, len(col)) + 2
                    worksheet.set_column(col_num + (1 if is_totals else 0), col_num + (1 if is_totals else 0),
                                         column_len)
                if is_totals:
                    worksheet.set_column(0, 0, max(df.index.astype(str).str.len().max(), len('Office')) + 2)

                worksheet.autofilter(0, 0, len(df), len(df.columns) + (1 if is_totals else 0) - 1)

            license_counts_worksheet = writer.sheets['License Counts']
            format_sheet(license_counts_worksheet, license_counts_df, is_totals=True)

            for col_num in range(len(license_counts_df.columns)):
                license_counts_worksheet.write(len(license_counts_df), col_num + 1, license_counts_df.iloc[-1, col_num],
                                               total_format)
            license_counts_worksheet.write(len(license_counts_df), 0, 'Total', total_format)

            # Set currency format for cost columns
            user_cost_col = 'Cost of Users (${})'.format(cost_per_user)
            exchange_cost_col = 'Cost of Exchange Licenses (${})'.format(cost_per_exchange)
            e5_cost_col = 'Cost of E5 Licenses (${})'.format(cost_per_e5)
            teams_cost_col = 'Cost of Teams Licenses (${})'.format(cost_per_teams)
            billable_total_col = 'Billable Total'

            user_cost_idx = license_counts_df.columns.get_loc(user_cost_col) + 1
            exchange_cost_idx = license_counts_df.columns.get_loc(exchange_cost_col) + 1
            e5_cost_idx = license_counts_df.columns.get_loc(e5_cost_col) + 1
            teams_cost_idx = license_counts_df.columns.get_loc(teams_cost_col) + 1
            billable_total_idx = license_counts_df.columns.get_loc(billable_total_col) + 1

            license_counts_worksheet.set_column(user_cost_idx, user_cost_idx, 15, currency_format)
            license_counts_worksheet.set_column(exchange_cost_idx, exchange_cost_idx, 15, currency_format)
            license_counts_worksheet.set_column(e5_cost_idx, e5_cost_idx, 15, currency_format)
            license_counts_worksheet.set_column(teams_cost_idx, teams_cost_idx, 15, currency_format)
            license_counts_worksheet.set_column(billable_total_idx, billable_total_idx, 15, currency_format)

            aion_management_worksheet = writer.sheets['AION Management']
            format_sheet(aion_management_worksheet, aion_management_df)

            aion_partners_worksheet = writer.sheets['AION Partners']
            format_sheet(aion_partners_worksheet, aion_partners_df)

            properties_worksheet = writer.sheets['Properties']
            format_sheet(properties_worksheet, properties_df)

            unaccounted_users_worksheet = writer.sheets['Unaccounted Users']
            format_sheet(unaccounted_users_worksheet, unaccounted_users)
        logger.info(f"Data saved to Excel file: {excel_path}")
    except PermissionError:
        logger.error(f"Permission denied when writing to {excel_path}")
    except Exception as e:
        logger.error(f"Error writing to Excel file {excel_path}: {e}")


def process_file(file_path, cost_per_user=115, cost_per_exchange=20, cost_per_e5=54.80, cost_per_teams=4):
    """
        Main function to process the CSV file and generate the Excel report.

        Args:
            file_path (str): Path to the input CSV file.
            cost_per_user (int, optional): Cost per user. Defaults to 115.
            cost_per_exchange (int, optional): Cost per exchange license. Defaults to 20.
            cost_per_e5 (int, optional): Cost per E5 license. Defaults to 300.
            cost_per_teams (int, optional): Cost per Teams license. Defaults to 10.

        Returns:
            str or None: Path to the generated Excel file if successful, None otherwise.
        """
    logger.info(f"Processing file: {file_path}")
    start_time = time.time()

    # Sanitize input path
    try:
        base_dir = os.path.dirname(os.path.abspath(file_path))
        file_path = sanitize_path(base_dir, os.path.basename(file_path))
    except ValueError as e:
        logger.error(f"Invalid file path: {e}")
        return None

    df = read_and_prepare_data(file_path)
    if df is None:
        return None

    target_licenses = {
        '365 Premium': ['Microsoft 365 Business Premium', 'E3'],
        'Exchange': ['Exchange'],
        'E5': ['E5'],
        'Teams': ['Microsoft Teams Enterprise']
    }

    license_counts = initialize_license_counts(df, target_licenses)
    license_counts, unaccounted_users, aion_management, aion_partners, properties = process_licenses(df,
                                                                                                     target_licenses,
                                                                                                     license_counts)

    license_counts_df = create_license_counts_df(license_counts)
    if license_counts_df is None:
        return None

    license_counts_df['Cost of Users (${})'.format(cost_per_user)] = license_counts_df['365 Premium'] * cost_per_user
    license_counts_df['Cost of Exchange Licenses (${})'.format(cost_per_exchange)] = license_counts_df[
                                                                                         'Exchange'] * cost_per_exchange
    license_counts_df['Cost of E5 Licenses (${})'.format(cost_per_e5)] = license_counts_df['E5'] * cost_per_e5
    license_counts_df['Cost of Teams Licenses (${})'.format(cost_per_teams)] = license_counts_df[
                                                                                   'Teams'] * cost_per_teams
    license_counts_df['Billable Total'] = (
            license_counts_df['Cost of Users (${})'.format(cost_per_user)] +
            license_counts_df['Cost of Exchange Licenses (${})'.format(cost_per_exchange)] +
            license_counts_df['Cost of E5 Licenses (${})'.format(cost_per_e5)] +
            license_counts_df['Cost of Teams Licenses (${})'.format(cost_per_teams)]
    )

    aion_management_df = pd.DataFrame(aion_management, columns=['Display Name', 'License Type', 'User Principal Name'])
    aion_partners_df = pd.DataFrame(aion_partners, columns=['Display Name', 'License Type', 'User Principal Name'])
    properties_df = pd.DataFrame(properties,
                                 columns=['Display Name', 'License Type', 'User Principal Name', 'Office'])
    unaccounted_users = pd.DataFrame(unaccounted_users, columns=['Display Name', 'License Type', 'User Principal Name'])

    current_date = datetime.now().strftime('%Y_%m_%d')
    file_id = str(uuid.uuid4())
    internal_filename = f"{file_id}_license_counts_{current_date}.xlsx"
    friendly_filename = f"AION_License_Report_{current_date}.xlsx"
    excel_path = os.path.join(app.config['OUTPUT_FOLDER'], internal_filename)

    save_to_excel(excel_path, license_counts_df, aion_management_df, aion_partners_df, properties_df, unaccounted_users,
                  cost_per_user,
                  cost_per_exchange, cost_per_e5, cost_per_teams)
    logger.info(f"Processed file saved to: {excel_path}")

    try:
        os.remove(file_path)
        logger.info(f"Removed uploaded file: {file_path}")
    except OSError as e:
        logger.error(f"Error removing uploaded {file_path}: {e}")

    end_time = time.time()
    processing_time = end_time - start_time
    logger.info(f"CSV processing time: {processing_time:.2f} seconds")

    return excel_path, friendly_filename


def generate_summary(file_path):
    """
        Generate a summary of the license counts from the Excel file.

        Args:
            file_path (str): Path to the Excel file.

        Returns:
            dict: Summary statistics of the license counts.
        """
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['License Counts']

    total_row = sheet.max_row
    data = [
        [sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1)]
        for i in range(2, total_row)  # Skip header row
    ]

    summary = {
        'total_365_premium': sum(row[1] for row in data),
        'total_exchange': sum(row[2] for row in data),
        'total_e5': sum(row[3] for row in data),
        'total_teams': sum(row[4] for row in data),
        'total_cost': sum(row[6] for row in data),
        'avg_cost_per_office': sum(row[6] for row in data) / len(data),
        'highest_cost': max(row[6] for row in data),
        'highest_cost_office': next(row[0] for row in data if row[6] == max(r[6] for r in data)),
        'percent_with_e5': sum(1 for row in data if row[3] > 0) / len(data) * 100,
        'offices_with_e5': sum(1 for row in data if row[3] > 0),
        'percent_both_licenses': sum(1 for row in data if row[1] > 0 and row[2] > 0) / len(data) * 100,
        'percent_only_365': sum(1 for row in data if row[1] > 0 and row[2] == 0 and row[3] == 0 and row[4] == 0) / len(
            data) * 100,
        'percent_only_exchange': sum(
            1 for row in data if row[1] == 0 and row[2] > 0 and row[3] == 0 and row[4] == 0) / len(data) * 100,
        'percent_only_e5': sum(1 for row in data if row[1] == 0 and row[2] == 0 and row[3] > 0 and row[4] == 0) / len(
            data) * 100,
        'percent_only_teams': sum(
            1 for row in data if row[1] == 0 and row[2] == 0 and row[3] == 0 and row[4] > 0) / len(data) * 100,
        'highest_exchange_ratio': max((row[2] / row[1] if row[1] > 0 else 0) for row in data),
        'highest_exchange_ratio_office': next(row[0] for row in data if (row[2] / row[1] if row[1] > 0 else 0) == max(
            (r[2] / r[1] if r[1] > 0 else 0) for r in data)),
        'highest_e5_ratio': max((row[3] / (row[1] + row[2]) if (row[1] + row[2]) > 0 else 0) for row in data),
        'highest_e5_ratio_office': next(row[0] for row in data if
                                           (row[3] / (row[1] + row[2]) if (row[1] + row[2]) > 0 else 0) == max(
                                               (r[3] / (r[1] + r[2]) if (r[1] + r[2]) > 0 else 0) for r in data)),
        'highest_teams_ratio': max(
            (row[4] / (row[1] + row[2] + row[3]) if (row[1] + row[2] + row[3]) > 0 else 0) for row in data),
        'highest_teams_ratio_office': next(row[0] for row in data if
                                           (row[4] / (row[1] + row[2] + row[3]) if (row[1] + row[2] + row[
                                               3]) > 0 else 0) == max(
                                               (r[4] / (r[1] + r[2] + r[3]) if (r[1] + r[2] + r[3]) > 0 else 0) for r in
                                               data)),
        'offices_no_licenses': sum(1 for row in data if row[1] == 0 and row[2] == 0 and row[3] == 0 and row[4] == 0),
        'avg_licenses_per_office': (sum(row[1] for row in data) + sum(row[2] for row in data) +
                                    sum(row[3] for row in data) + sum(row[4] for row in data)) / len(data),
        'top_offices_by_license': sorted(
            [(row[0], row[1], row[2], row[3], row[4], row[1] + row[2] + row[3] + row[4]) for row in data],
            key=lambda x: x[5],
            reverse=True
        )[:5],
    }

    return summary
